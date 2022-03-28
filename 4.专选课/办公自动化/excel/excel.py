# -*- coding:utf-8 -*-
'''
@File    :   excel.py
@Time    :   2022/03/28 08:05:06
@Author  :   hyh
@Version :   1.0
@Contact :   1360895771@qq.com
@Desc    :   自动办公,课上练习
'''
# here put the import lib
import openpyxl

#wb = openpyxl.load_workbook('7.公选课\\办公自动化\\example.xlsx')
wb = openpyxl.load_workbook('example.xlsx')
wb.sheetnames
sheet = wb['Sheet1']
print(sheet['A1'].value)

c = sheet['B1']
print(c.value)

print(sheet['C1'].value)

for i in range(1,5):
    print(i,sheet.cell(i,2).value)#cell根据行列查找单元隔，均从1开始

print(sheet.max_column,sheet.max_row)

#输出矩形区域
for row in sheet['A1':'C3']:
    for cellobj in row:
        print(cellobj.coordinate,cellobj.value)
    print('---END OF ROW---')

#按列输出单元格信息
for i in sheet.rows:
    print(i)

#统计各位同学的日常开销
print("下面是各位同学的日常开销:")
# sum=0
# for rowobj in sheet['B1':'C3']:
#     if rowobj[0].vlaue == 'Li Lei':
#         sum = sum+rowobj[1].value
#         print(sum)
sh = wb.active#选中正在活动的sheet
pay_sum = { }
for row in sh.rows:
    fields = [c.value for c in row]
    name = fields[1]
    pay = fields[2]
    pay_sum[name] = pay_sum.get(name,0) + pay
print(pay_sum)
import pprint
resultFile = open('result.py','w')
resultFile.write(
    'pay_sum = '+pprint.pformat(pay_sum)
)
resultFile.close()

#创建和删除工作簿或表
wb1 = openpyxl.Workbook()
print(wb1.sheetnames)
sh1 = wb1.active
sh1.title = "My Data"
wb1.save("mydata.xlsx")

#单元格操作（写入、属性设置等）

#小九九
wb2 = openpyxl.Workbook()
sheet = wb2.active
for i in range(1,10):
    for j in range(1,10):
        sheet.cell(row = i+1,column= j+1).value = i*j
wb2.save("九九乘法表.xlsx")

#把result.py文件存储到xlsx中
import result
row=6
for k,v in result.pay_sum.items():
    sh.cell(row=row,column=1).value = k
    sh.cell(row=row,column=2).value = v
    row += 1